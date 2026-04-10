"""Étape 6 — Export Excel + Word fusionné."""
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import streamlit as st

from core.db import get_dossier, charger_adresses, changer_statut, mettre_a_jour_parametres
from core.validator import a_alerte_bloquante


st.title("Export final")
st.caption("Étape 6 / 6")

dossier_id = st.session_state.get("dossier_id")
if not dossier_id:
    st.switch_page("app.py")
    st.stop()

dossier = get_dossier(dossier_id)
st.markdown(f"**Dossier :** {dossier['nom']}")

adresses = st.session_state.get("adresses")
if not adresses:
    with st.spinner("Chargement depuis la base..."):
        adresses = charger_adresses(dossier_id)
if not adresses:
    st.error("Aucune adresse. Reprenez depuis Composition.")
    st.stop()

# ---------------------------------------------------------------------------
# Export Excel
# ---------------------------------------------------------------------------

FILL_ROUGE = PatternFill("solid", fgColor="FCEAEA")
FILL_ORANGE = PatternFill("solid", fgColor="FEF3E6")

EN_TETES = ["ID_CLIENT", "Formule", "L1", "L2", "L3", "L4", "L5", "L6"]


def generer_excel(adresses: list[dict]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Adresses_NormAdress"

    # En-têtes
    for col, h in enumerate(EN_TETES, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row_idx, a in enumerate(adresses, 2):
        # Résoudre clés Supabase (l1→L1) ou composer (L1)
        def get(k):
            return a.get(k) or a.get(k.lower()) or ""

        alertes = a.get("alertes", [])
        if isinstance(alertes, str):
            import json
            alertes = json.loads(alertes)

        bloquant = a_alerte_bloquante(alertes) if alertes else False
        fill = FILL_ROUGE if bloquant else (FILL_ORANGE if alertes else None)

        vals = [
            get("id_client"),
            get("Formule") or get("formule"),
            get("L1"), get("L2"), get("L3"),
            get("L4"), get("L5"), get("L6"),
        ]
        for col, val in enumerate(vals, 1):
            cell = ws.cell(row=row_idx, column=col, value=val or "")
            if fill:
                cell.fill = fill
            # CP en texte (colonne L6 — index 8)
            if col == 8 and val:
                cell.number_format = "@"

    # Largeurs automatiques
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


st.markdown("### Export Excel")
if st.button("Générer l'Excel", type="primary"):
    with st.spinner("Génération Excel..."):
        excel_bytes = generer_excel(adresses)
    st.download_button(
        label="Télécharger adresses_normadress.xlsx",
        data=excel_bytes,
        file_name="adresses_normadress.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ---------------------------------------------------------------------------
# Tips publipostage Word
# ---------------------------------------------------------------------------

st.markdown("---")
st.markdown("### Publipostage Word")
with st.expander("📋 Comment réaliser le publipostage dans Word"):
    st.markdown("""
1. Ouvrir Word → onglet **Publipostage** → **Démarrer la fusion et le publipostage** → **Lettres**
2. **Sélection des destinataires** → **Utiliser une liste existante** → choisir l'Excel exporté (onglet *Adresses_NormAdress*)
3. Positionner le curseur à l'emplacement de l'adresse dans le courrier
4. **Insérer un champ de fusion** pour chaque ligne : `L1`, `L2`, `L3`, `L4`, `L5`, `L6`
5. ⚠️ **Lignes vides** — pour éviter les lignes blanches quand L2/L3/L5 sont vides :
   - Sélectionner le champ + son saut de ligne
   - **Règles** → **Si… Alors… Sinon…** : Si *champ* = **(vide)** → ne rien afficher
6. Insérer également le champ `Formule` pour la salutation personnalisée
7. **Terminer et fusionner** → **Modifier des documents individuels** → **Tous** → OK
8. Word génère un nouveau document avec toutes les lettres — imprimer ou exporter en PDF
""")

# ---------------------------------------------------------------------------
# Clôture du dossier
# ---------------------------------------------------------------------------

st.markdown("---")
statut_export = dossier["statut"]
if statut_export != "exporte":
    if st.button("Marquer le dossier comme exporté ✅", type="primary"):
        params = dossier.get("parametres") or {}
        params["date_export"] = datetime.now().strftime("%d/%m/%Y %H:%M")
        mettre_a_jour_parametres(dossier_id, params)
        changer_statut(dossier_id, "exporte")
        st.success("Dossier clôturé.")
        st.rerun()

# Recul d'état
with st.expander("⚙️ Modifier l'état du dossier"):
    st.caption("Permet de revenir en arrière si le client demande une modification.")
    if st.button("↩ Remettre à valider", use_container_width=True):
        changer_statut(dossier_id, "valide")
        st.rerun()
    if st.button("↩ Remettre en cours", use_container_width=True):
        changer_statut(dossier_id, "en_cours")
        st.rerun()
