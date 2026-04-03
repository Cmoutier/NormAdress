"""Génération du fichier Excel propre et du rapport texte."""
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import pandas as pd

from .mapper import STANDARD_FIELDS

# Couleurs
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
HEADER_FILL = PatternFill(start_color="1E6B3C", end_color="1E6B3C", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def export_excel(
    df: pd.DataFrame,
    original_df: pd.DataFrame,
    rapport_lignes: list[dict],
    doublons: list[dict],
    consolidation_journal: list[dict],
) -> bytes:
    """
    Génère un fichier Excel avec coloration :
    - Vert : cellule corrigée par rapport à l'original
    - Orange : consolidation d'adresse
    - Rouge : doublon
    Retourne les bytes du fichier.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Données nettoyées"

    # En-têtes
    for col_idx, field in enumerate(STANDARD_FIELDS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=field)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Ensemble des lignes consolidées (index 0-based dans df)
    consolidated_rows = {entry["ligne"] - 2 for entry in consolidation_journal}

    # Ensemble des lignes en doublon
    doublon_row_numbers = set()
    for d in doublons:
        for ligne_num in d["lignes"]:
            doublon_row_numbers.add(ligne_num - 2)  # index 0-based

    # Données
    orig_cols = set(original_df.columns) if original_df is not None else set()

    for row_idx, (df_idx, row) in enumerate(df.iterrows(), start=2):
        is_doublon = row_idx - 2 in doublon_row_numbers
        is_consolidated = row_idx - 2 in consolidated_rows

        for col_idx, field in enumerate(STANDARD_FIELDS, start=1):
            value = row.get(field, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)

            if is_doublon:
                cell.fill = RED_FILL
            elif is_consolidated and field.startswith("Adresse"):
                cell.fill = ORANGE_FILL
            elif original_df is not None and field in orig_cols:
                try:
                    orig_val = str(original_df.at[df_idx, field]) if df_idx in original_df.index else None
                except (KeyError, IndexError):
                    orig_val = None
                if orig_val is not None and str(value) != orig_val:
                    cell.fill = GREEN_FILL

    # Ajustement largeur colonnes
    for col_idx in range(1, len(STANDARD_FIELDS) + 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def export_rapport(
    nb_importees: int,
    nb_exportees: int,
    nb_supprimees: int,
    rapport_lignes: list[dict],
    doublons: list[dict],
    consolidation_journal: list[dict],
) -> str:
    """Génère un rapport texte complet."""
    lines = []
    lines.append("=" * 60)
    lines.append("RAPPORT DE MISE EN CONFORMITÉ — NormAdress")
    lines.append(f"Date : {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    lines.append("=" * 60)
    lines.append("")
    lines.append("RÉSUMÉ")
    lines.append(f"  Lignes importées  : {nb_importees}")
    lines.append(f"  Lignes exportées  : {nb_exportees}")
    lines.append(f"  Lignes supprimées : {nb_supprimees}")
    lines.append(f"  Doublons détectés : {len(doublons)}")
    lines.append(f"  Consolidations    : {len(consolidation_journal)}")
    lines.append(f"  Anomalies         : {len(rapport_lignes)}")
    lines.append("")

    if consolidation_journal:
        lines.append("-" * 60)
        lines.append("CONSOLIDATIONS D'ADRESSE")
        for entry in consolidation_journal:
            lines.append(f"  Ligne {entry['ligne']} :")
            lines.append(f"    Avant : {entry['avant']}")
            lines.append(f"    Après : {entry['apres']}")
        lines.append("")

    if doublons:
        lines.append("-" * 60)
        lines.append("DOUBLONS DÉTECTÉS (non supprimés)")
        for d in doublons:
            lignes_str = ", ".join(str(l) for l in d["lignes"])
            lines.append(f"  {d['nom']} / {d['codepostal']} → lignes {lignes_str}")
        lines.append("")

    if rapport_lignes:
        lines.append("-" * 60)
        lines.append("ANOMALIES ET AVERTISSEMENTS")
        for entry in rapport_lignes:
            lines.append(f"  Ligne {entry['ligne']} [{entry['colonne']}] : {entry['message']}")
        lines.append("")

    lines.append("=" * 60)
    lines.append("Fin du rapport")
    return "\n".join(lines)
