"""Tests pour cleaner/exporter.py."""
import io
import pytest
import pandas as pd
import openpyxl
from cleaner.exporter import export_excel, export_rapport
from cleaner.mapper import STANDARD_FIELDS


def make_result_df(rows=None):
    if rows is None:
        rows = [
            {"Civilite": "M.", "Nom": "DUPONT", "Prenom": "Jean",
             "Societe": "", "Adresse1": "10 rue de la Paix",
             "Adresse2": "", "Adresse3": "", "CodePostal": "75001", "Ville": "PARIS"},
        ]
    return pd.DataFrame(rows, columns=STANDARD_FIELDS)


class TestExportExcel:
    def test_returns_bytes(self):
        df = make_result_df()
        result = export_excel(df, df.copy(), [], [], [])
        assert isinstance(result, bytes)
        assert len(result) > 0

    def test_valid_xlsx(self):
        df = make_result_df()
        raw = export_excel(df, df.copy(), [], [], [])
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        assert len(wb.sheetnames) >= 1

    def test_headers_present(self):
        df = make_result_df()
        raw = export_excel(df, df.copy(), [], [], [])
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb.active
        headers = [ws.cell(row=1, column=i+1).value for i in range(len(STANDARD_FIELDS))]
        assert headers == STANDARD_FIELDS

    def test_data_rows(self):
        df = make_result_df()
        raw = export_excel(df, df.copy(), [], [], [])
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb.active
        assert ws.cell(row=2, column=2).value == "DUPONT"

    def test_consolidation_coloring(self):
        df = make_result_df()
        journal = [{"ligne": 2, "avant": {}, "apres": {}, "message": "", "type": "consolidation"}]
        raw = export_excel(df, df.copy(), [], [], journal)
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        # Juste vérifier que ça ne plante pas
        assert wb is not None

    def test_doublon_coloring(self):
        df = make_result_df([
            {"Civilite": "M.", "Nom": "DUPONT", "Prenom": "Jean", "Societe": "",
             "Adresse1": "10 rue", "Adresse2": "", "Adresse3": "",
             "CodePostal": "75001", "Ville": "PARIS"},
            {"Civilite": "M.", "Nom": "DUPONT", "Prenom": "Jean", "Societe": "",
             "Adresse1": "10 rue", "Adresse2": "", "Adresse3": "",
             "CodePostal": "75001", "Ville": "PARIS"},
        ])
        doublons = [{"cle": "dupont|75001", "lignes": [2, 3], "nom": "DUPONT", "codepostal": "75001"}]
        raw = export_excel(df, df.copy(), [], doublons, [])
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        assert wb is not None

    def test_without_original_df(self):
        df = make_result_df()
        raw = export_excel(df, None, [], [], [])
        assert isinstance(raw, bytes)

    def test_multiple_rows(self):
        df = make_result_df([
            {"Civilite": "M.", "Nom": f"NOM{i}", "Prenom": f"Prenom{i}", "Societe": "",
             "Adresse1": f"{i} rue", "Adresse2": "", "Adresse3": "",
             "CodePostal": f"7500{i}", "Ville": "PARIS"}
            for i in range(5)
        ])
        raw = export_excel(df, df.copy(), [], [], [])
        wb = openpyxl.load_workbook(io.BytesIO(raw))
        ws = wb.active
        assert ws.max_row == 6  # 1 en-tête + 5 données


class TestExportRapport:
    def test_returns_string(self):
        result = export_rapport(10, 9, 1, [], [], [])
        assert isinstance(result, str)

    def test_contains_summary(self):
        result = export_rapport(100, 95, 5, [], [], [])
        assert "100" in result
        assert "95" in result
        assert "5" in result

    def test_contains_doublons(self):
        doublons = [{"nom": "DUPONT", "codepostal": "75001", "lignes": [2, 5]}]
        result = export_rapport(10, 10, 0, [], doublons, [])
        assert "DUPONT" in result
        assert "75001" in result
        assert "2" in result
        assert "5" in result

    def test_contains_consolidations(self):
        journal = [{"ligne": 3, "avant": {"Adresse1": ""}, "apres": {"Adresse1": "10 rue"}, "message": ""}]
        result = export_rapport(10, 10, 0, [], [], journal)
        assert "Ligne 3" in result
        assert "consolidation" in result.lower()

    def test_contains_anomalies(self):
        rapport = [{"ligne": 4, "colonne": "CodePostal", "message": "invalide", "type": "error"}]
        result = export_rapport(10, 10, 0, rapport, [], [])
        assert "CodePostal" in result
        assert "invalide" in result

    def test_no_doublons_section_when_empty(self):
        result = export_rapport(10, 10, 0, [], [], [])
        assert "DOUBLONS" not in result

    def test_no_consolidation_section_when_empty(self):
        result = export_rapport(10, 10, 0, [], [], [])
        assert "CONSOLIDATION" not in result

    def test_encoding_utf8(self):
        result = export_rapport(1, 1, 0, [], [], [])
        encoded = result.encode("utf-8")
        assert len(encoded) > 0

    def test_date_present(self):
        result = export_rapport(1, 1, 0, [], [], [])
        # Le rapport contient une date au format DD/MM/YYYY
        import re
        assert re.search(r"\d{2}/\d{2}/\d{4}", result)
