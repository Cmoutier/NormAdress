"""Script de génération des fixtures de test."""
import openpyxl
import os

FIXTURES_DIR = os.path.join(os.path.dirname(__file__), "fixtures")
os.makedirs(FIXTURES_DIR, exist_ok=True)


def make_demo_standard():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Feuille1"
    ws.append(["Civilite", "Nom", "Prenom", "Societe", "Adresse1", "Adresse2", "Adresse3", "CodePostal", "Ville"])
    ws.append(["Mr", "dupont", "jean-pierre", "ACME Corp", "10 rue de la Paix", "", "", "75001", "paris"])
    ws.append(["Mme", "MARTIN", "marie", "", "5 avenue Victor Hugo", "Apt 3B", "", "69000", "LYON"])
    ws.append(["Mlle", "bernard", "sophie", "Tech SA", "", "", "", "13001", "marseille"])
    ws.append(["", "leclerc", "paul", "", "20 rue du Commerce", "", "", "1000", "bordeaux"])
    wb.save(os.path.join(FIXTURES_DIR, "demo_standard.xlsx"))


def make_demo_multi_contacts():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contacts"
    ws.append(["Civilite1", "Nom1", "Prenom1", "Civilite2", "Nom2", "Prenom2",
               "Adresse1", "CodePostal", "Ville"])
    ws.append(["Mr", "dupont", "jean", "Mme", "dupont", "marie",
               "10 rue de la Paix", "75001", "paris"])
    ws.append(["Mr", "martin", "paul", "Mr", "martin", "pierre",
               "5 avenue Foch", "69000", "lyon"])
    wb.save(os.path.join(FIXTURES_DIR, "demo_multi_contacts.xlsx"))


def make_demo_adresses_mal_remplies():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Adresses"
    ws.append(["Nom", "Adresse1", "Adresse2", "Adresse3", "CodePostal", "Ville"])
    ws.append(["dupont", "", "10 rue de la Paix", "Bat A", "75001", "paris"])
    ws.append(["martin", "", "", "5 avenue Victor Hugo", "69000", "lyon"])
    ws.append(["bernard", "20 rue du Commerce", "", "", "13001", "marseille"])
    wb.save(os.path.join(FIXTURES_DIR, "demo_adresses_mal_remplies.xlsx"))


def make_demo_colonnes_synonymes():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Export"
    ws.append(["lastname", "firstname", "city", "zip", "street"])
    ws.append(["dupont", "jean", "paris", "75001", "10 rue de la Paix"])
    ws.append(["martin", "marie", "lyon", "69000", "5 avenue Victor Hugo"])
    wb.save(os.path.join(FIXTURES_DIR, "demo_colonnes_synonymes.xlsx"))


def make_demo_csv():
    content = "Civilite;Nom;Prenom;Adresse1;CodePostal;Ville\nMr;dupont;jean;10 rue de la Paix;75001;paris\nMme;martin;marie;5 avenue Victor Hugo;69000;lyon\n"
    with open(os.path.join(FIXTURES_DIR, "demo_csv_semicolon.csv"), "w", encoding="utf-8") as f:
        f.write(content)


if __name__ == "__main__":
    make_demo_standard()
    make_demo_multi_contacts()
    make_demo_adresses_mal_remplies()
    make_demo_colonnes_synonymes()
    make_demo_csv()
    print("Fixtures créées dans", FIXTURES_DIR)
